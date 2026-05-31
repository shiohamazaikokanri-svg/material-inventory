from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from .models import MaterialItem


HEADER_ALIASES = {
    "key": {"key", "id", "item key", "品番", "キー"},
    "maker": {"maker", "manufacturer", "brand", "メーカー"},
    "material": {"material", "材質", "材料"},
    "thickness": {"thickness", "thick", "t", "板厚", "厚み"},
    "width": {"width", "w", "幅", "巾"},
    "length": {"length", "l", "長さ"},
    "quantity": {"quantity", "qty", "count", "数量", "枚数"},
    "weight": {"weight", "wt", "重量"},
    "place": {"place", "location", "storage", "場所"},
    "address": {"address", "棚番", "住所", "保管先", "番地"},
}

REQUIRED_FIELDS = {"key", "material", "thickness", "width", "length", "quantity", "weight"}


def import_materials_from_excel(uploaded_file):
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0, ["Excelファイルが空です。"]

    header_map = _build_header_map(rows[0])
    missing = sorted(REQUIRED_FIELDS - set(header_map))
    if missing:
        return 0, [f"必須列が見つかりません: {', '.join(missing)}"]

    created = 0
    errors = []
    items = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(value in (None, "") for value in row):
            continue
        try:
            items.append(
                MaterialItem(
                    key=_text(row, header_map, "key"),
                    maker=_text(row, header_map, "maker"),
                    material=_text(row, header_map, "material"),
                    thickness=_decimal(row, header_map, "thickness"),
                    width=_decimal(row, header_map, "width"),
                    length=_decimal(row, header_map, "length"),
                    quantity=_int(row, header_map, "quantity"),
                    weight=_decimal(row, header_map, "weight"),
                    place=_text(row, header_map, "place"),
                    address=_text(row, header_map, "address"),
                )
            )
        except ValueError as exc:
            errors.append(f"Row {row_number}: {exc}")

    if items:
        MaterialItem.objects.bulk_create(items, batch_size=500)
        created = len(items)

    return created, errors


def _build_header_map(header_row):
    normalized = {_normalize(value): index for index, value in enumerate(header_row) if value is not None}
    header_map = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                header_map[field] = normalized[alias]
                break
    return header_map


def _normalize(value):
    return str(value).strip().lower()


def _cell(row, header_map, field):
    index = header_map.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def _text(row, header_map, field):
    value = _cell(row, header_map, field)
    return "" if value is None else str(value).strip()


def _decimal(row, header_map, field):
    value = _cell(row, header_map, field)
    if value in (None, ""):
        raise ValueError(f"{field} は必須です")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError(f"{field} は数値で入力してください") from exc


def _int(row, header_map, field):
    value = _decimal(row, header_map, field)
    if value < 0:
        raise ValueError(f"{field} はマイナスにできません")
    return int(value)
